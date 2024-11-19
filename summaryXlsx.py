from openpyxl import load_workbook, Workbook
import glob,os
import datetime


orgi_path = "D:/00 dinglidan/业务/18 测试活动/10版本测试"
#orgi_path = "D:/00 dinglidan/业务/18 测试活动/10版本测试/SOW3架构整改测试/02 Mirror"
srcfilekeyword = "测试用例"
sheetkeyword = "统计"
headRow = 3  # 表头有多少行 或者 表头到第几行结束
keycol = 3   # 名称或者编号等关键的列
colkeyword = "特性"  # 关键列的名称，用于检测表单是否是需要汇总的

savapath ="D:/00 dinglidan/业务/18 测试活动/10版本测试/"
savefilename = "SOW3架构整改测试结果汇总"
current_time = datetime.datetime.now().strftime("%Y%m%d")
savefile = f"{savefilename}{current_time}.xlsx"

"""
    拷贝源单元格的全部，包括各种格式
"""
def copy_cell(source_cell, target_cell):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    # 复制值
    target_cell.value = source_cell.value
    # 复制数字格式（百分号等）
    target_cell.number_format = source_cell.number_format
    # 复制字体
    target_cell.font = Font(name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            vertAlign=source_cell.font.vertAlign,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            color=source_cell.font.color)

    # 复制填充
    target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                   fgColor=source_cell.fill.fgColor,
                                   bgColor=source_cell.fill.bgColor,
                                   patternType=source_cell.fill.patternType)

    # 复制边框
    border = Border(left=Side(border_style=source_cell.border.left.border_style,
                              color=source_cell.border.left.color),
                    right=Side(border_style=source_cell.border.right.border_style,
                               color=source_cell.border.right.color),
                    top=Side(border_style=source_cell.border.top.border_style,
                             color=source_cell.border.top.color),
                    bottom=Side(border_style=source_cell.border.bottom.border_style,
                                color=source_cell.border.bottom.color))
    target_cell.border = border

    # 复制对齐方式
    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      text_rotation=source_cell.alignment.textRotation,
                                      wrap_text=source_cell.alignment.wrapText,
                                      shrink_to_fit=source_cell.alignment.shrinkToFit,
                                      indent=source_cell.alignment.indent)

    # 复制保护
    target_cell.protection = Protection(locked=source_cell.protection.locked,
                                        hidden=source_cell.protection.hidden)


"""
    判断一行是否为空行。空行的定义是所有单元格的值都为空。
"""
def is_empty_row(row):
    for cell in row:
        if cell.value is not None:
            return False
    return True


"""
    删除工作表中的空行。
"""
def delete_empty_rows(sheet):
    # 记录需要删除的行
    rows_to_delete = []
    # 从最后一行开始遍历，这样删除行不会影响未处理的行索引
    for row in range(sheet.max_row, 0, -1):
        if is_empty_row(sheet[row]):
            rows_to_delete.append(row)
    print("要删除的行有——", rows_to_delete)

    # 删除记录的空行
    for row in rows_to_delete:
        sheet.delete_rows(row)


"""
    目录下每个文件，先检查关键字，再拷贝表头（一次），最后计算数据行的行号，拷贝每个文件数据；删除空行，保存
"""

# 新建汇总表
total_workbook = Workbook()
total_sheet = total_workbook.active
head_flag = 0  # 头只拷贝一次
data_row = 0   # 汇总表中，数据行累加，记录每个文件开始拷贝位置
print("新建汇总表---ok")

# 递归遍历子目录，并筛选文件名包含“srcfilekeyword”的文件
excel_files = glob.glob(os.path.join(orgi_path, '**', '*.xlsx'), recursive=True)
filtered_files = [file for file in excel_files if srcfilekeyword in os.path.basename(file)]  # ding

for file in filtered_files:
    # 遍历每个源文件
    orig_workbook = load_workbook(file, data_only=True)
    try:
        orig_sheet = orig_workbook[sheetkeyword]
    except KeyError as e:
        print(f"KeyError: {e}")   # ding
        continue # 跳过错误文件，汇总下一个文件
    print("待拷贝的文件", file)

    # 检查文件内容
    key_value = orig_sheet.cell(row=headRow, column=keycol).value
    if key_value != colkeyword:
        print("文件错误", headRow, keycol, key_value)
        continue   # 跳过错误文件，汇总下一个文件

    # 拷贝表头
    if not head_flag:
        for hRow in range(1, headRow+1):
            # 拷贝表头的一行
            for headcell in orig_sheet[hRow]:
                target_cell = total_sheet.cell(row=headcell.row, column=headcell.column)
                copy_cell(headcell, target_cell)
            print("拷贝表头", hRow)
        # 复制合并单元格
        for merge_range in orig_sheet.merged_cells.ranges:  # ding
            total_sheet.merge_cells(merge_range.coord)
        # 数据行开始
        data_row = headRow + 1
        head_flag = 1

    # 找需要拷贝的数据行，"keycol"列是关键列； ”headRow“下一行是数据行
    rowNolst = []
    for col in orig_sheet.iter_cols(min_col=keycol, max_col=keycol, min_row=headRow+1, max_row=orig_sheet.max_row):
        for cell in col:
            rowNolst.append(cell.row)
    print("数据行的行号有：", rowNolst, len(rowNolst))

    # 拷贝数据
    for row in rowNolst:
        print("数据准备拷贝行：", data_row)
        # 拷贝一行
        for datacell in orig_sheet[row]:
            target_cell = total_sheet.cell(row=data_row, column=datacell.column) # 目标单元格与源单元格列同，行累加
            copy_cell(datacell, target_cell)
        data_row = data_row + 1
    print("数据已存好-done")


delete_empty_rows(total_sheet)

total_workbook.save(savapath + savefile)
print("汇总完成，保持在这里——", savapath + savefile)


